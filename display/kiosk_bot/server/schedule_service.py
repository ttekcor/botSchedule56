import os
from openpyxl import load_workbook

UPLOAD_FOLDER = os.path.join(os.getcwd(), 'botSchedule56', 'display', 'kiosk_bot', 'server', 'upload')

def get_schedule_from_excel(file_name):
    """Читает файл Excel и возвращает данные в формате JSON."""
    file_path = os.path.join(UPLOAD_FOLDER, file_name)
    workbook = load_workbook(file_path)
    sheet = workbook.active

    schedule = []
    for row in sheet.iter_rows(values_only=True):
        schedule.append(list(row))

    return schedule

def get_teacher_schedule(file_name, teacher_name=None):
    """Возвращает расписание для конкретного учителя или всех учителей."""
    file_path = os.path.join(UPLOAD_FOLDER, file_name)
    workbook = load_workbook(file_path)
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))

    # Получаем названия классов, начиная со второго столбца, через каждые три колонки
    class_names = [data[0][i] for i in range(2, len(data[0]), 3)]

    teachers = {}

    # Парсим каждую строку данных, начиная со второй, где расписание уроков
    for row in data[1:]:
        for i, class_name in enumerate(class_names):
            teacher1 = row[2 + i * 3]
            teacher2 = row[3 + i * 3] if 3 + i * 3 < len(row) else None
            lesson_number = row[0]
            lesson_name = row[1 + i * 3]

            if teacher1:
                if teacher1 not in teachers:
                    teachers[teacher1] = []
                teachers[teacher1].append({'number': lesson_number, 'lesson': lesson_name, 'class': class_name})

            if teacher2:
                if teacher2 not in teachers:
                    teachers[teacher2] = []
                teachers[teacher2].append({'number': lesson_number, 'lesson': lesson_name, 'class': class_name})

    if teacher_name:
        return teachers.get(teacher_name)

    return [{'teacher': teacher, 'schedule': schedule} for teacher, schedule in teachers.items()]
